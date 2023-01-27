import sys
import openpyxl
from io import StringIO

def main():
    xlsx = sys.argv[1]
    wb = openpyxl.load_workbook(filename = xlsx)

    file = StringIO()
    file.write('<?xml version="1.0" encoding="utf-8"?>\n')
    file.write('<Book Count="{}">\n'.format(len(wb.sheetnames)/2))
    for name in wb.sheetnames:
        if not name.endswith('Header'):
            header = wb[name + 'Header']
            sheet = wb[name]
            file.write('\t<Sheet Name="{name}" Count="{count}">\n'.format(name = name, count = sheet.max_row - 1))

            file.write('\t\t<Header>\n')
            for row in range(2, header.max_row + 1):
                param = ''
                param += '\t\t\t<Param '
                for column in range(1, header.max_column + 1):
                    field = header.cell(row = row, column = column).value
                    if (isinstance(field, str)):
                        field = field.replace('&', '&amp;')
                        field = field.replace('>', '&gt;')
                        field = field.replace('<', '&lt;')
                        field = field.replace('"', '&quot;')
                    param += '{key}="{value}" '.format(key = header.cell(row = 1, column = column).value, value = field)
                param += '/>\n'
                file.write(param)
            file.write('\t\t</Header>\n')
            
            file.write('\t\t<Data>\n')
            for row in range(2, sheet.max_row + 1):
                param = ''
                param += '\t\t\t<Param '
                for column in range(1, sheet.max_column + 1):
                    field = sheet.cell(row = row, column = column).value
                    if (isinstance(field, str)):
                        field = field.replace('&', '&amp;')
                        field = field.replace('>', '&gt;')
                        field = field.replace('<', '&lt;')
                        field = field.replace('"', '&quot;')
                    param += '{key}="{value}" '.format(key = sheet.cell(row = 1, column = column).value, value = field)
                param += '/>\n'
                file.write(param)
            file.write('\t\t</Data>\n')
            file.write('\t</Sheet>\n')
    file.write('</Book>')
    xml = open(sys.argv[2], mode = 'w', encoding='utf8')
    xml.write(file.getvalue())

if __name__ == "__main__":
    main()
